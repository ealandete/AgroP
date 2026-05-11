from fastapi import APIRouter, Depends, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
from datetime import date, datetime
import io, csv
from app.database import get_db
from app.models import (
    Animal, Siembra, Venta, Costo, Ordeno, Pesaje, Sanidad,
    Usuario, Lote, Producto, Finca, Reproduccion, Lactancia,
    AlimentacionDiaria, MovimientoAnimal, PlanificacionPastoreo, LaborCampo,
    Raza, VariedadCultivo, Cosecha,
)
from app.utils.auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["Exportacion"])


def get_model(model_name: str):
    models = {
        "animales": Animal, "siembras": Siembra, "ventas": Venta,
        "costos": Costo, "ordenos": Ordeno, "pesajes": Pesaje,
        "sanidad": Sanidad, "usuarios": Usuario, "lotes": Lote,
        "productos": Producto, "reproduccion": Reproduccion,
        "lactancias": Lactancia, "alimentacion": AlimentacionDiaria,
        "movimientos": MovimientoAnimal, "pastoreo": PlanificacionPastoreo,
        "labores": LaborCampo,
    }
    return models.get(model_name)


def model_to_dict(obj):
    d = {}
    for c in obj.__table__.columns:
        val = getattr(obj, c.name)
        if isinstance(val, (date,)):
            val = str(val)
        d[c.name] = val
    return d


def get_finca_name(db):
    f = db.query(Finca).first()
    return f.nombre if f else "AgroP"


# ── PDF helpers ──────────────────────────────────────────

def _pdf_header_footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("Helvetica", 7)
    canvas.setFillColorRGB(0.4, 0.4, 0.4)
    canvas.drawString(40, 20, f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    canvas.drawRightString(letter[0] - 40, 20, f"Pagina {doc.page}")
    canvas.restoreState()


def _build_pdf(elements, title="Reporte"):
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, PageTemplate, Frame

    buf = io.BytesIO()
    frame = Frame(40, 40, letter[0] - 80, letter[1] - 80, id="normal")
    template = PageTemplate(id="main", frames=[frame], onPage=_pdf_header_footer)
    doc = SimpleDocTemplate(buf, pagesize=letter)
    doc.addPageTemplates([template])
    doc.build(elements)
    buf.seek(0)
    return buf


def _make_table(data, col_widths=None):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet

    t = Table(data, colWidths=col_widths, repeatRows=1)
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]
    t.setStyle(TableStyle(style))
    return t


def _heading(text, level=1):
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    styles = getSampleStyleSheet()
    if level == 1:
        s = styles["Title"]
    elif level == 2:
        s = styles["Heading2"]
    else:
        s = styles["Heading3"]
    from reportlab.platypus import Paragraph
    return Paragraph(text, s)


def _paragraph(text):
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph
    return Paragraph(text, getSampleStyleSheet()["Normal"])


def _spacer(h=12):
    from reportlab.platypus import Spacer
    return Spacer(1, h)


from reportlab.lib.pagesizes import letter


# ── CSV export ───────────────────────────────────────────

@router.get("/csv/{model_name}")
def export_csv(
    model_name: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    Model = get_model(model_name)
    if not Model:
        return {"error": "Modelo no encontrado"}

    records = db.query(Model).all()
    if not records:
        return {"error": "Sin datos"}

    output = io.StringIO()
    headers = [c.name for c in Model.__table__.columns]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for r in records:
        writer.writerow(model_to_dict(r))

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={model_name}.csv"},
    )


# ── Excel export ─────────────────────────────────────────

@router.get("/excel/{model_name}")
def export_excel(
    model_name: str,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    Model = get_model(model_name)
    if not Model:
        return {"error": "Modelo no encontrado"}

    records = db.query(Model).all()
    if not records:
        return {"error": "Sin datos"}

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = model_name

    headers = [c.name for c in Model.__table__.columns]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row, r in enumerate(records, 2):
        d = model_to_dict(r)
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=d.get(h))

    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={model_name}.xlsx"},
    )


# ── PDF Reporte Animales ─────────────────────────────────

@router.get("/pdf/reporte-animales")
def export_pdf_animales(
    especie: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak

    q = db.query(Animal).filter(Animal.activo == True)
    if especie:
        q = q.filter(Animal.especie == especie)
    animales = q.all()
    razas = {r.id: r.nombre for r in db.query(Raza).all()}
    finca_nombre = get_finca_name(db)

    elements = []
    elements.append(_heading(f"Reporte de Animales - {finca_nombre}", 1))
    elements.append(_paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"))
    elements.append(_spacer(8))

    # Group by especie
    from collections import defaultdict
    grupos = defaultdict(list)
    for a in animales:
        grupos[a.especie or "Sin especie"].append(a)

    total_general = len(animales)
    for especie_nombre, grupo in sorted(grupos.items()):
        elements.append(_heading(f"{especie_nombre.capitalize()} ({len(grupo)})", 2))
        data = [["Codigo", "Nombre", "Raza", "Sexo", "Peso (kg)", "Lote", "Estado"]]
        for a in grupo:
            data.append([
                a.codigo or "", a.nombre or "", razas.get(a.raza_id) or "",
                a.sexo or "", str(float(a.peso_kg)) if a.peso_kg else "-",
                a.lote.nombre if a.lote else "-",
                "Activo" if a.activo else "Inactivo",
            ])
        elements.append(_make_table(data, col_widths=[60, 70, 70, 40, 55, 70, 50]))
        elements.append(_spacer(6))

    elements.append(_spacer(10))
    elements.append(_paragraph(f"<b>Total animales: {total_general}</b>"))

    buf = _build_pdf(elements, "reporte_animales")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_animales.pdf"},
    )


# ── PDF Reporte Cultivos ─────────────────────────────────

@router.get("/pdf/reporte-cultivos")
def export_pdf_cultivos(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from reportlab.platypus import Paragraph, Spacer

    siembras = db.query(Siembra).filter(Siembra.estado == "activo").all()
    lotes = {l.id: l.nombre for l in db.query(Lote).all()}
    variedades = {v.id: v.variedad for v in db.query(VariedadCultivo).all()}
    finca_nombre = get_finca_name(db)

    elements = []
    elements.append(_heading(f"Reporte de Cultivos - {finca_nombre}", 1))
    elements.append(_paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"))
    elements.append(_spacer(8))

    data = [["Cultivo", "Variedad", "Lote", "Fecha Siembra", "Area (ha)", "Rendimiento (kg/ha)", "Estado"]]
    total_area = 0
    for s in siembras:
        total_area += float(s.area_ha or 0)
        data.append([
            s.cultivo.capitalize() if s.cultivo else "",
            variedades.get(s.variedad_id) or "-",
            lotes.get(s.lote_id) or "-",
            str(s.fecha_siembra) if s.fecha_siembra else "",
            str(s.area_ha) if s.area_ha else "-",
            str(float(s.rendimiento_ha)) if s.rendimiento_ha else "-",
            s.estado or "",
        ])

    elements.append(_make_table(data, col_widths=[60, 60, 70, 65, 55, 70, 55]))
    elements.append(_spacer(8))
    elements.append(_paragraph(f"<b>Siembras activas: {len(siembras)} | Area total: {total_area:.2f} ha</b>"))

    # Also show cosechas
    cosechas = db.query(Cosecha).all()
    if cosechas:
        elements.append(_spacer(12))
        elements.append(_heading("Resumen de Cosechas", 2))
        data_c = [["Fecha", "Cantidad (kg)", "Calidad", "Destino"]]
        total_cosecha = 0
        for c in cosechas:
            total_cosecha += float(c.cantidad_kg or 0)
            data_c.append([str(c.fecha) if c.fecha else "", str(float(c.cantidad_kg)) if c.cantidad_kg else "-", c.calidad or "-", c.destino or "-"])
        elements.append(_make_table(data_c, col_widths=[70, 80, 70, 90]))
        elements.append(_spacer(4))
        elements.append(_paragraph(f"<b>Total cosechado: {total_cosecha:.2f} kg</b>"))

    buf = _build_pdf(elements, "reporte_cultivos")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_cultivos.pdf"},
    )


# ── PDF Reporte Financiero ──────────────────────────────

@router.get("/pdf/reporte-financiero")
def export_pdf_financiero(
    desde: Optional[date] = None,
    hasta: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from reportlab.platypus import Paragraph, Spacer

    qv = db.query(Venta)
    qc = db.query(Costo)
    if desde:
        qv = qv.filter(Venta.fecha >= desde)
        qc = qc.filter(Costo.fecha >= desde)
    if hasta:
        qv = qv.filter(Venta.fecha <= hasta)
        qc = qc.filter(Costo.fecha <= hasta)

    ventas = qv.order_by(Venta.fecha.desc()).all()
    costos = qc.order_by(Costo.fecha.desc()).all()
    total_ingresos = sum(float(v.total or 0) for v in ventas)
    total_gastos = sum(float(c.monto or 0) for c in costos)
    balance = total_ingresos - total_gastos
    finca_nombre = get_finca_name(db)

    elements = []
    elements.append(_heading(f"Reporte Financiero - {finca_nombre}", 1))
    periodo = f"{desde or 'Inicio'} - {hasta or 'Hoy'}"
    elements.append(_paragraph(f"Periodo: {periodo} | Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"))
    elements.append(_spacer(8))

    # Summary
    data_resumen = [
        ["Concepto", "Valor"],
        ["Total Ingresos", f"${total_ingresos:,.0f}"],
        ["Total Gastos", f"${total_gastos:,.0f}"],
        ["Balance Neto", f"${balance:,.0f}"],
    ]
    elements.append(_heading("Estado de Resultados", 2))
    elements.append(_make_table(data_resumen, col_widths=[200, 150]))
    elements.append(_spacer(12))

    # Ventas detail
    elements.append(_heading(f"Ventas ({len(ventas)})", 2))
    if ventas:
        data_v = [["Fecha", "Producto", "Cliente", "Cantidad", "Total"]]
        for v in ventas:
            prod = db.query(Producto).filter(Producto.id == v.producto_id).first()
            data_v.append([
                str(v.fecha), prod.nombre if prod else "", v.cliente or "",
                str(float(v.cantidad)) if v.cantidad else "-",
                f"${float(v.total or 0):,.0f}",
            ])
        elements.append(_make_table(data_v, col_widths=[65, 100, 80, 55, 80]))
    else:
        elements.append(_paragraph("Sin ventas en el periodo."))

    elements.append(_spacer(12))

    # Gastos detail
    elements.append(_heading(f"Gastos ({len(costos)})", 2))
    if costos:
        data_c = [["Fecha", "Descripcion", "Categoria", "Monto"]]
        for c in costos:
            cat_nombre = c.categoria.nombre if c.categoria else "-"
            data_c.append([
                str(c.fecha), c.descripcion or "", cat_nombre,
                f"${float(c.monto or 0):,.0f}",
            ])
        elements.append(_make_table(data_c, col_widths=[65, 140, 80, 80]))
    else:
        elements.append(_paragraph("Sin gastos en el periodo."))

    buf = _build_pdf(elements, "reporte_financiero")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_financiero.pdf"},
    )


# ── PDF Reporte Lotes ───────────────────────────────────

@router.get("/pdf/reporte-lotes")
def export_pdf_lotes(
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from reportlab.platypus import Paragraph, Spacer

    lotes = db.query(Lote).filter(Lote.activo == True).all()
    siembras = db.query(Siembra).filter(Siembra.estado == "activo").all()
    siembras_por_lote = {}
    for s in siembras:
        siembras_por_lote.setdefault(s.lote_id, []).append(s)
    finca_nombre = get_finca_name(db)

    elements = []
    elements.append(_heading(f"Reporte de Lotes - {finca_nombre}", 1))
    elements.append(_paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"))
    elements.append(_spacer(8))

    data = [["Nombre", "Codigo", "Area (ha)", "Uso Actual", "Suelo", "Riego", "Cultivos"]]
    total_area = 0
    for l in lotes:
        total_area += float(l.area_ha or 0)
        cults = siembras_por_lote.get(l.id, [])
        cultivos_str = ", ".join(set(s.cultivo for s in cults)) if cults else "-"
        data.append([
            l.nombre or "", l.codigo or "",
            str(float(l.area_ha)) if l.area_ha else "-",
            l.uso_actual or "-", l.tipo_suelo or "-",
            l.sistema_riego or "-", cultivos_str,
        ])

    elements.append(_make_table(data, col_widths=[65, 50, 50, 60, 55, 60, 90]))
    elements.append(_spacer(8))
    elements.append(_paragraph(f"<b>Total lotes: {len(lotes)} | Area total: {total_area:.2f} ha</b>"))

    buf = _build_pdf(elements, "reporte_lotes")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_lotes.pdf"},
    )


# ── PDF Reporte Completo ────────────────────────────────

@router.get("/pdf/reporte-completo")
def export_pdf_completo(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from reportlab.platypus import Paragraph, Spacer, PageBreak
    from collections import defaultdict

    finca_nombre = get_finca_name(db)
    razas = {r.id: r.nombre for r in db.query(Raza).all()}
    lotes = {l.id: l.nombre for l in db.query(Lote).all()}

    elements = []
    elements.append(_heading(f"Reporte Completo de Finca - {finca_nombre}", 1))
    elements.append(_paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"))
    elements.append(_spacer(12))

    # ── Animales ──
    elements.append(_heading("Inventario de Animales", 2))
    animales = db.query(Animal).filter(Animal.activo == True).all()
    grupos = defaultdict(list)
    for a in animales:
        grupos[a.especie or "Sin especie"].append(a)
    for esp, grupo in sorted(grupos.items()):
        data = [["Codigo", "Nombre", "Raza", "Sexo", "Peso"]]
        for a in grupo:
            data.append([a.codigo or "", a.nombre or "", razas.get(a.raza_id) or "", a.sexo or "", str(float(a.peso_kg)) if a.peso_kg else "-"])
        elements.append(_paragraph(f"<b>{esp.capitalize()} ({len(grupo)})</b>"))
        elements.append(_make_table(data, col_widths=[55, 65, 65, 35, 45]))
        elements.append(_spacer(4))
    elements.append(_paragraph(f"<b>Total animales: {len(animales)}</b>"))
    elements.append(PageBreak())

    # ── Cultivos ──
    elements.append(_heading("Cultivos Activos", 2))
    siembras = db.query(Siembra).filter(Siembra.estado == "activo").all()
    data_c = [["Cultivo", "Lote", "Fecha Siembra", "Area (ha)", "Rendimiento"]]
    total_area = 0
    for s in siembras:
        total_area += float(s.area_ha or 0)
        data_c.append([
            s.cultivo.capitalize() if s.cultivo else "", lotes.get(s.lote_id) or "-",
            str(s.fecha_siembra) if s.fecha_siembra else "",
            str(s.area_ha) if s.area_ha else "-",
            str(float(s.rendimiento_ha)) if s.rendimiento_ha else "-",
        ])
    elements.append(_make_table(data_c, col_widths=[60, 65, 65, 50, 60]))
    elements.append(_paragraph(f"<b>Siembras activas: {len(siembras)} | Area: {total_area:.2f} ha</b>"))

    cosechas = db.query(Cosecha).all()
    if cosechas:
        elements.append(_spacer(8))
        elements.append(_paragraph("<b>Cosechas</b>"))
        data_co = [["Fecha", "Cantidad (kg)", "Calidad"]]
        total_cosecha = sum(float(c.cantidad_kg or 0) for c in cosechas)
        for c in cosechas[:10]:
            data_co.append([str(c.fecha), str(float(c.cantidad_kg)), c.calidad or "-"])
        elements.append(_make_table(data_co, col_widths=[65, 70, 60]))
        elements.append(_paragraph(f"Total cosechado: {total_cosecha:.2f} kg"))
    elements.append(PageBreak())

    # ── Lotes ──
    elements.append(_heading("Lotes y Terrenos", 2))
    lotes_data = db.query(Lote).filter(Lote.activo == True).all()
    data_l = [["Nombre", "Area (ha)", "Uso", "Suelo", "Riego"]]
    total_area_l = 0
    for l in lotes_data:
        total_area_l += float(l.area_ha or 0)
        data_l.append([l.nombre or "", str(float(l.area_ha)) if l.area_ha else "-", l.uso_actual or "-", l.tipo_suelo or "-", l.sistema_riego or "-"])
    elements.append(_make_table(data_l, col_widths=[60, 50, 60, 55, 60]))
    elements.append(_paragraph(f"<b>Total lotes: {len(lotes_data)} | Area: {total_area_l:.2f} ha</b>"))
    elements.append(PageBreak())

    # ── Financiero ──
    qv2 = db.query(Venta)
    qc2 = db.query(Costo)
    if desde:
        qv2 = qv2.filter(Venta.fecha >= desde)
        qc2 = qc2.filter(Costo.fecha >= desde)
    if hasta:
        qv2 = qv2.filter(Venta.fecha <= hasta)
        qc2 = qc2.filter(Costo.fecha <= hasta)
    ventas_f = qv2.all()
    costos_f = qc2.all()
    ti = sum(float(v.total or 0) for v in ventas_f)
    tg = sum(float(c.monto or 0) for c in costos_f)

    elements.append(_heading("Resumen Financiero", 2))
    periodo = f"{desde or 'Inicio'} - {hasta or 'Hoy'}"
    elements.append(_paragraph(f"Periodo: {periodo}"))
    data_f = [
        ["Concepto", "Valor"],
        ["Ingresos", f"${ti:,.0f}"],
        ["Gastos", f"${tg:,.0f}"],
        ["Balance", f"${ti - tg:,.0f}"],
    ]
    elements.append(_make_table(data_f, col_widths=[200, 150]))

    buf = _build_pdf(elements, "reporte_completo")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_completo.pdf"},
    )


# ── Excel Reporte Completo ──────────────────────────────

@router.get("/excel/reporte-completo")
def export_excel_completo(
    desde: Optional[date] = Query(None),
    hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_sheet(ws, title, headers, rows):
        ws.title = title
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align
            c.border = thin_border
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = thin_border
        ws.auto_filter.ref = ws.dimensions

    # Sheet 1: Animales
    ws1 = wb.active
    animales = db.query(Animal).filter(Animal.activo == True).all()
    razas = {r.id: r.nombre for r in db.query(Raza).all()}
    rows_a = []
    for a in animales:
        rows_a.append([
            a.codigo or "", a.nombre or "", a.especie or "",
            razas.get(a.raza_id) or "", a.sexo or "",
            str(a.fecha_nacimiento) if a.fecha_nacimiento else "",
            float(a.peso_kg) if a.peso_kg else "",
            a.lote.nombre if a.lote else "",
            "Activo" if a.activo else "Inactivo",
        ])
    write_sheet(ws1, "Animales",
        ["Codigo", "Nombre", "Especie", "Raza", "Sexo", "F. Nacimiento", "Peso (kg)", "Lote", "Estado"],
        rows_a)

    # Sheet 2: Cultivos
    ws2 = wb.create_sheet()
    siembras = db.query(Siembra).filter(Siembra.estado == "activo").all()
    lotes_map = {l.id: l.nombre for l in db.query(Lote).all()}
    variedades = {v.id: v.variedad for v in db.query(VariedadCultivo).all()}
    rows_c = []
    for s in siembras:
        rows_c.append([
            s.cultivo.capitalize() if s.cultivo else "",
            variedades.get(s.variedad_id) or "",
            lotes_map.get(s.lote_id) or "",
            str(s.fecha_siembra) if s.fecha_siembra else "",
            float(s.area_ha) if s.area_ha else "",
            float(s.cantidad_semilla) if s.cantidad_semilla else "",
            float(s.rendimiento_ha) if s.rendimiento_ha else "",
            s.estado or "",
        ])
    write_sheet(ws2, "Cultivos",
        ["Cultivo", "Variedad", "Lote", "Fecha Siembra", "Area (ha)", "Semilla", "Rendimiento (kg/ha)", "Estado"],
        rows_c)

    # Sheet 3: Lotes
    ws3 = wb.create_sheet()
    lotes_q = db.query(Lote).filter(Lote.activo == True).all()
    rows_l = []
    for l in lotes_q:
        rows_l.append([
            l.nombre or "", l.codigo or "", float(l.area_ha) if l.area_ha else "",
            l.uso_actual or "", l.tipo_suelo or "", l.sistema_riego or "",
            float(l.latitud) if l.latitud else "",
            float(l.longitud) if l.longitud else "",
        ])
    write_sheet(ws3, "Lotes",
        ["Nombre", "Codigo", "Area (ha)", "Uso Actual", "Suelo", "Riego", "Latitud", "Longitud"],
        rows_l)

    # Sheet 4: Financiero
    ws4 = wb.create_sheet()
    qv = db.query(Venta)
    qc = db.query(Costo)
    if desde:
        qv = qv.filter(Venta.fecha >= desde)
        qc = qc.filter(Costo.fecha >= desde)
    if hasta:
        qv = qv.filter(Venta.fecha <= hasta)
        qc = qc.filter(Costo.fecha <= hasta)
    ventas = qv.all()
    costos = qc.all()
    rows_f = []
    for v in ventas:
        prod = db.query(Producto).filter(Producto.id == v.producto_id).first()
        rows_f.append([
            "Ingreso", str(v.fecha), prod.nombre if prod else "", v.cliente or "",
            float(v.cantidad), float(v.total or 0),
        ])
    for c in costos:
        rows_f.append([
            "Gasto", str(c.fecha), c.descripcion or "", "",
            "", float(c.monto or 0),
        ])
    write_sheet(ws4, "Financiero",
        ["Tipo", "Fecha", "Concepto", "Cliente", "Cantidad", "Valor"],
        rows_f)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_completo.xlsx"},
    )


# ── Excel Import ─────────────────────────────────────────

@router.post("/importar/animales")
def importar_animales(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from openpyxl import load_workbook

    errors = []
    created = 0

    try:
        wb = load_workbook(file.file, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(min_row=2, values_only=True)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    except Exception as e:
        return {"success": False, "error": f"Error leyendo archivo: {str(e)}", "created": 0, "errors": []}

    finca_id = current_user.finca_id or 1

    for row_idx, row in enumerate(rows_iter, start=2):
        if not any(row):
            continue
        try:
            row_dict = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                row_dict[h] = val

            codigo = str(row_dict.get("codigo") or row_dict.get("Codigo") or "")
            nombre = str(row_dict.get("nombre") or row_dict.get("Nombre") or "")
            especie = str(row_dict.get("especie") or row_dict.get("Especie") or "bovino")
            sexo = str(row_dict.get("sexo") or row_dict.get("Sexo") or "H")
            peso = row_dict.get("peso_kg") or row_dict.get("Peso (kg)") or None
            raza_nombre = str(row_dict.get("raza") or row_dict.get("Raza") or "")

            raza_id = None
            if raza_nombre:
                raza = db.query(Raza).filter(Raza.nombre == raza_nombre, Raza.especie == especie).first()
                if raza:
                    raza_id = raza.id

            animal = Animal(
                finca_id=finca_id,
                codigo=codigo or None,
                nombre=nombre or None,
                especie=especie,
                sexo=sexo if sexo in ("M", "H") else "H",
                peso_kg=float(peso) if peso else None,
                raza_id=raza_id,
                fecha_ingreso=date.today(),
                activo=True,
            )
            db.add(animal)
            created += 1

        except Exception as e:
            errors.append({"fila": row_idx, "error": str(e)})

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"success": False, "error": f"Error al guardar: {str(e)}", "created": created, "errors": errors}

    return {"success": True, "created": created, "errors": errors}


@router.post("/importar/siembras")
def importar_siembras(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: Usuario = Depends(get_current_user),
):
    from openpyxl import load_workbook

    errors = []
    created = 0

    try:
        wb = load_workbook(file.file, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(min_row=2, values_only=True)
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    except Exception as e:
        return {"success": False, "error": f"Error leyendo archivo: {str(e)}", "created": 0, "errors": []}

    for row_idx, row in enumerate(rows_iter, start=2):
        if not any(row):
            continue
        try:
            row_dict = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                row_dict[h] = val

            cultivo = str(row_dict.get("cultivo") or row_dict.get("Cultivo") or "")
            lote_nombre = str(row_dict.get("lote") or row_dict.get("Lote") or "")
            fecha_str = str(row_dict.get("fecha_siembra") or row_dict.get("Fecha Siembra") or "")
            area = row_dict.get("area_ha") or row_dict.get("Area (ha)") or None

            if not cultivo:
                errors.append({"fila": row_idx, "error": "Cultivo es requerido"})
                continue

            lote_id = None
            if lote_nombre:
                lote = db.query(Lote).filter(Lote.nombre == lote_nombre).first()
                if lote:
                    lote_id = lote.id

            fecha_siembra = date.today()
            if fecha_str:
                try:
                    fecha_siembra = datetime.strptime(fecha_str, "%Y-%m-%d").date()
                except ValueError:
                    try:
                        fecha_siembra = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                    except ValueError:
                        pass

            siembra = Siembra(
                cultivo=cultivo,
                lote_id=lote_id,
                fecha_siembra=fecha_siembra,
                area_ha=float(area) if area else None,
                estado="activo",
            )
            db.add(siembra)
            created += 1

        except Exception as e:
            errors.append({"fila": row_idx, "error": str(e)})

    try:
        db.commit()
    except Exception as e:
        db.rollback()
        return {"success": False, "error": f"Error al guardar: {str(e)}", "created": created, "errors": errors}

    return {"success": True, "created": created, "errors": errors}
