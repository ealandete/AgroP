import io
from datetime import date, datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import Animal, Siembra, Lote, Insumo, Inventario, GrupoManejo, Finca, Raza
from app.utils.auth import get_current_user

router = APIRouter(prefix="/api/templates", tags=["Plantillas"])

TEMPLATE_CONFIGS = {
    "animales": {
        "columns": [
            "codigo", "nombre", "especie", "sexo", "raza",
            "fecha_nacimiento", "peso_kg", "color", "lote_id", "estado_origen",
        ],
        "sample": [
            "ANM-001", "Ejemplo", "bovino", "H", "Holstein",
            "2023-05-15", "350.5", "Negro con blanco", "1", "propio",
        ],
        "validations": {
            "sexo": {"type": "list", "formula": '"M,H"'},
            "especie": {"type": "list", "formula": '"bovino,bufalino,porcino,aviar,ovino,caprino,equino"'},
            "estado_origen": {"type": "list", "formula": '"propio,prestamo,adopcion,consignacion"'},
        },
    },
    "siembras": {
        "columns": [
            "cultivo", "variedad", "lote_id", "fecha_siembra",
            "area_ha", "metodo_siembra",
        ],
        "sample": [
            "maiz", "ICA V-305", "1", "2024-01-15", "2.5", "mecanica",
        ],
        "validations": {
            "metodo_siembra": {"type": "list", "formula": '"manual,mecanica,siembra_directa,trasplante"'},
        },
    },
    "lotes": {
        "columns": [
            "nombre", "codigo", "area_ha", "tipo_suelo",
            "uso_actual", "latitud", "longitud",
        ],
        "sample": [
            "Potrero Norte", "L-001", "10.5", "franco_arenoso",
            "pastoreo", "6.2447", "-75.5748",
        ],
    },
    "insumos": {
        "columns": [
            "nombre", "tipo", "unidad_medida", "categoria", "stock_minimo",
        ],
        "sample": [
            "Fertilizante NPK 15-15-15", "fertilizante", "kg",
            "Fertilizantes", "100",
        ],
    },
    "inventario": {
        "columns": [
            "insumo_codigo", "cantidad", "costo_unitario",
            "fecha_ingreso", "ubicacion",
        ],
        "sample": [
            "INS-001", "500", "2500.00", "2024-01-10", "Bodega Central",
        ],
        "validations": {},
    },
}

UPLOAD_HISTORY = []


def _style_header(ws, num_cols):
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _apply_validations(ws, config, num_rows):
    validations = config.get("validations", {})
    columns = config["columns"]
    for col_name, v in validations.items():
        if col_name in columns:
            col_idx = columns.index(col_name) + 1
            col_letter = get_column_letter(col_idx)
            from openpyxl.worksheet.datavalidation import DataValidation
            dv = DataValidation(
                type=v["type"],
                formula1=v["formula"],
                allow_blank=True,
            )
            dv.error = "Valor no valido"
            dv.errorTitle = "Error de validacion"
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}{num_rows}")


def _auto_width(ws, num_cols, columns):
    for col in range(1, num_cols + 1):
        max_len = max(len(str(columns[col - 1])), 14)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)


@router.post("/upload/{tipo}")
async def cargar_plantilla(tipo: str, file: UploadFile = File(...), db: Session = Depends(get_db)):
    config = TEMPLATE_CONFIGS.get(tipo)
    if not config:
        raise HTTPException(status_code=404, detail=f"Tipo de plantilla no valido: {tipo}")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    content = await file.read()
    try:
        wb = Workbook()
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    columns = config["columns"]
    errors = []
    created = 0

    for row_idx, row in enumerate(rows, 2):
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            continue
        row_data = {}
        for col_idx, col_name in enumerate(columns):
            val = row[col_idx] if col_idx < len(row) else None
            if isinstance(val, str):
                val = val.strip()
            if hasattr(val, "isoformat"):
                val = val.isoformat()
            row_data[col_name] = val

        try:
            if tipo == "animales":
                _procesar_animal(row_data, db)
            elif tipo == "siembras":
                _procesar_siembra(row_data, db)
            elif tipo == "lotes":
                _procesar_lote(row_data, db)
            elif tipo == "insumos":
                _procesar_insumo(row_data, db)
            elif tipo == "inventario":
                _procesar_inventario(row_data, db)
            created += 1
        except Exception as e:
            errors.append(f"Fila {row_idx}: {str(e)}")

    if created > 0:
        db.commit()

    record = {
        "timestamp": datetime.now().isoformat(),
        "tipo": tipo,
        "filename": file.filename,
        "creados": created,
        "errores": errors,
    }
    UPLOAD_HISTORY.append(record)

    return {
        "message": f"Procesados {created} registros con {len(errors)} errores",
        "creados": created,
        "errores": errors,
        "total_filas": len(rows),
    }


def _procesar_animal(data: dict, db: Session):
    finca_id = _get_finca_id(db)
    raza_id = None
    if data.get("raza"):
        raza = db.query(Raza).filter(
            Raza.nombre == data["raza"]
        ).first()
        if raza:
            raza_id = raza.id

    lote_id = None
    if data.get("lote_id"):
        try:
            lote_id = int(data["lote_id"])
        except (ValueError, TypeError):
            pass

    animal = Animal(
        finca_id=finca_id,
        codigo=data.get("codigo"),
        nombre=data.get("nombre"),
        especie=data.get("especie", "bovino"),
        sexo=data.get("sexo", "H"),
        raza_id=raza_id,
        fecha_nacimiento=_parse_date(data.get("fecha_nacimiento")),
        peso_kg=_parse_float(data.get("peso_kg")),
        color=data.get("color"),
        lote_id=lote_id,
        estado_origen=data.get("estado_origen", "propio"),
        fecha_ingreso=date.today(),
    )
    db.add(animal)


def _procesar_siembra(data: dict, db: Session):
    finca_id = _get_finca_id(db)
    lote_id = int(data.get("lote_id", 0))
    lote = db.query(Lote).filter(Lote.id == lote_id, Lote.finca_id == finca_id).first()
    if not lote:
        raise ValueError(f"Lote {lote_id} no encontrado")

    siembra = Siembra(
        lote_id=lote_id,
        cultivo=data.get("cultivo", ""),
        variedad=_resolve_variedad(data.get("variedad"), data.get("cultivo")),
        fecha_siembra=_parse_date(data.get("fecha_siembra")),
        area_ha=_parse_float(data.get("area_ha")),
        metodo_siembra=data.get("metodo_siembra"),
        estado="activo",
    )
    db.add(siembra)


def _procesar_lote(data: dict, db: Session):
    finca_id = _get_finca_id(db)
    lote = Lote(
        finca_id=finca_id,
        nombre=data.get("nombre", ""),
        codigo=data.get("codigo"),
        area_ha=_parse_float(data.get("area_ha")),
        tipo_suelo=data.get("tipo_suelo"),
        uso_actual=data.get("uso_actual"),
        latitud=_parse_float(data.get("latitud")),
        longitud=_parse_float(data.get("longitud")),
    )
    db.add(lote)


def _procesar_insumo(data: dict, db: Session):
    insumo = Insumo(
        nombre=data.get("nombre", ""),
        tipo=data.get("tipo"),
        unidad_medida=data.get("unidad_medida", "unidad"),
        stock_minimo=_parse_float(data.get("stock_minimo")),
    )
    db.add(insumo)


def _procesar_inventario(data: dict, db: Session):
    finca_id = _get_finca_id(db)
    insumo_codigo = data.get("insumo_codigo")
    insumo = db.query(Insumo).filter(Insumo.codigo == insumo_codigo).first()
    if not insumo:
        raise ValueError(f"Insumo con codigo '{insumo_codigo}' no encontrado")

    inventario = Inventario(
        insumo_id=insumo.id,
        cantidad=_parse_float(data.get("cantidad", 0)),
        costo_unitario=_parse_float(data.get("costo_unitario")),
        fecha_ingreso=_parse_date(data.get("fecha_ingreso")),
        ubicacion=data.get("ubicacion"),
    )
    db.add(inventario)


def _get_finca_id(db: Session) -> int:
    finca = db.query(Finca).first()
    if not finca:
        raise HTTPException(status_code=400, detail="No hay fincas registradas")
    return finca.id


def _parse_date(val):
    if val is None:
        return date.today()
    if isinstance(val, date):
        return val
    try:
        return datetime.strptime(str(val), "%Y-%m-%d").date()
    except ValueError:
        return date.today()


def _parse_float(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _resolve_variedad(variedad_str, cultivo):
    if not variedad_str:
        return variedad_str
    return variedad_str


@router.get("/google-drive/auth-url")
def google_drive_auth_url():
    return {
        "auth_url": "https://accounts.google.com/o/oauth2/auth?mock=true&redirect_uri=urn:ietf:wg:oauth:2.0:oob",
        "mensaje": "URL de autorizacion simulada. En produccion se integraria con Google Drive API.",
    }


@router.post("/google-drive/import")
async def google_drive_import(fileId: str = "", tipo: str = "animales"):
    if not fileId:
        raise HTTPException(status_code=400, detail="fileId es requerido")

    config = TEMPLATE_CONFIGS.get(tipo)
    if not config:
        raise HTTPException(status_code=404, detail=f"Tipo de plantilla no valido: {tipo}")

    return {
        "message": f"Importacion desde Google Drive simulada para archivo {fileId}",
        "fileId": fileId,
        "tipo": tipo,
        "estado": "simulado",
        "nota": "En produccion se descargaria el archivo de Google Drive y se procesaria.",
    }


@router.get("/upload-history")
def upload_history():
    return UPLOAD_HISTORY[-50:]


@router.post("/upload/grupos-asignacion")
async def cargar_asignacion_grupos(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx o .xls")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content))
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo leer el archivo Excel")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    errors = []
    asignados = 0

    for row_idx, row in enumerate(rows, 2):
        if all(cell is None or (isinstance(cell, str) and cell.strip() == "") for cell in row):
            continue
        codigo = str(row[0]).strip() if row[0] else ""
        grupo_nombre = str(row[1]).strip() if len(row) > 1 and row[1] else ""

        if not codigo or not grupo_nombre:
            errors.append(f"Fila {row_idx}: codigo_animal y grupo_nombre son requeridos")
            continue

        try:
            animal = db.query(Animal).filter(Animal.codigo == codigo).first()
            if not animal:
                errors.append(f"Fila {row_idx}: Animal con codigo '{codigo}' no encontrado")
                continue

            grupo = db.query(GrupoManejo).filter(GrupoManejo.nombre == grupo_nombre).first()
            if not grupo:
                errors.append(f"Fila {row_idx}: Grupo '{grupo_nombre}' no encontrado")
                continue

            animal.grupo_manejo_id = grupo.id
            asignados += 1
        except Exception as e:
            errors.append(f"Fila {row_idx}: {str(e)}")

    if asignados > 0:
        db.commit()

    return {
        "message": f"Asignados {asignados} animales con {len(errors)} errores",
        "asignados": asignados,
        "errores": errors,
        "total_filas": len(rows),
    }


@router.get("/download/grupos-asignacion")
def descargar_plantilla_grupos():
    wb = Workbook()
    ws = wb.active
    ws.title = "Asignacion Grupos"

    columns = ["codigo_animal", "grupo_nombre"]
    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    _style_header(ws, len(columns))

    sample = ["ANM-001", "Grupo Sanitario"]
    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=2, column=col_idx, value=val)
        ws.cell(row=2, column=col_idx).font = Font(italic=True, color="666666")

    _auto_width(ws, len(columns), columns)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers = {
        "Content-Disposition": 'attachment; filename="plantilla_asignacion_grupos.xlsx"',
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.get("/{tipo}")
def descargar_plantilla(tipo: str):
    config = TEMPLATE_CONFIGS.get(tipo)
    if not config:
        raise HTTPException(status_code=404, detail=f"Tipo de plantilla no valido: {tipo}")

    wb = Workbook()
    ws = wb.active
    ws.title = tipo.capitalize()

    columns = config["columns"]
    sample = config["sample"]

    for col_idx, col_name in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    _style_header(ws, len(columns))

    for col_idx, val in enumerate(sample, 1):
        ws.cell(row=2, column=col_idx, value=val)
        cell = ws.cell(row=2, column=col_idx)
        cell.font = Font(italic=True, color="666666")

    _apply_validations(ws, config, 101)

    _auto_width(ws, len(columns), columns)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers = {
        "Content-Disposition": f'attachment; filename="plantilla_{tipo}.xlsx"',
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
